import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
import logging
import time

start_time = time.time()

# ****************** Load excel file into variable starts here *********************************
supplier_data = openpyxl.load_workbook("From supplier.xlsx")  # workbookname
system_data = openpyxl.load_workbook("From system.xlsx")  # workbookname

system_data_sheet = system_data["Catalog Lines"]  # sheet name

if "Catalog Lines" in supplier_data.sheetnames:
    supplier_data_sheet = supplier_data["Catalog Lines"]  # sheet name

elif "catalog lines" in supplier_data.sheetnames:
    supplier_data_sheet = supplier_data["catalog lines"]
else :
    supplier_data_sheet = supplier_data.active

# ------------------- Load excel file into variable ends here --------------


# ******************Defining colors starts here *********************************
Pattern_purple = PatternFill(patternType="solid", fgColor="ADD8E6")
Pattern_red = PatternFill(patternType="solid", fgColor="FFCCCB")
Pattern_yellow = PatternFill(patternType="solid", fgColor="ffff00")
Pattern_warning = PatternFill(patternType="solid", fgColor="FF0000")
# -----------------Defining colors end here ---------------------------

# *****************Making some list / variable object for future use*******************
list_of_row = []  # not yet consumed
list_of_SIN_from_supplier_data_sheet = []
list_of_UNSPSC = [420250001930, 420250001931, 420250001932, 420250001933, 420250001934, 420250001935, 420250001936, 420250001937, 420250001938,
                  420250001939, 420250001940, 420250001941, 420250001942, 420250001943, 420250001944, 420250001945, 420250001946, 420250001947,
                  420250001948, 420250001949, 420250001950, 420250001951, 420250001952, 420250001953, 420250001954, 420250001955, 420250001956,
                  420250001957, 420250001958, 420250001959, 420250001960, 420250001961, 420250001962, 420250001963, 420250001964, 420250001965,
                  420250001966, 420250001967, 420250001968, 420250001969, 420250001970, 420250001971, 420250001972, 420250001973, 420250001974,
                  420250001975, 420250001976, 420250001977, 420250001978, 420250001979, 420250001980, 420250001981, 420250001982, 420250001983,
                  420250001984, 420250001985, 420250001986, 420250001987, 420250001988, 420250001989, 420250001990, 420250001991, 420250001992,
                  420250001993, 420250001994, 420250001995, 420250001996, 420250001997, 420250001998, 420250001999, 420250002000, 420250002001,
                  420250002002, 420250002003, 420250002004, 420250002005, 420250002006, 420250002007, 420250002008, 420250002009, 420250002010,
                  420250002011, 420250002012, 420250002013, 420250002014, 420250002015, 420250002016, 420250002017, 420250002018, 420250002019,
                  420250002020, 420250002021, 420250002022, 420250002023, 420250002024, 420250002025, 420250002026, 420250002027, 420250002028,
                  420250002029, 420250002030, 420250002031, 420250002032, 420250002033, 420250002034, 420250002035, 420250002036, 420250002037,
                  420250002038, 420250002039, 420250002040, 420250002041, 420250002042, 420250002043, 420250002044, 420250002045, 420250002046,
                  420250002047, 420250002048, 420250002049, 420250002050, 420250002051, 420250002052, 420250002053, 420250002054, 420250002055,
                  420250002056, 420250002057, 420250002058, 420250002059, 420250002060, 420250002061, 420250002062, 420250002063, 420250002064,
                  420250002065, 420250002066, 420250002067, 420250002068, 420250002069, 420250002070, 420250002071, 420250002072, 420250002073,
                  420250002074, 420250002075, 420250002076, 420250002077, 420250002078, 420250002079, 420250002080, 420250002081, 420250002082,
                  420250002083, 420250002084, 420250002085, 420250002086, 420250002087, 420250002088, 420250002089, 420250002090, 420250002091,
                  420250002092, 420250002093, 420250002094, 420250002095, 420250002096, 420250002097, 420250002098, 420250002099, 420250002100,
                  420250002101, 420250002102, 420250002103, 420250002104, 420250002105, 420250002106, 420250002107, 420250002108, 420250002109,
                  420250002110, 420250002111, 420250002112, 420250002113, 420250002114, 420250002115, 420250002116, 420250002117, 420250002118,
                  420250002119, 420250002120, 420250002121, 420250002122, 420250002123, 420250002124, 420250002125, 420250002126, 420250002127,
                  420250002128, 420250002129, 420250002130, 420250002131, 420250002132, 420250002133, 420250002134, 420250002135, 420250002136,
                  420250002137, 420250002138, 420250002139, 420250002140, 420250002141, 420250002142, 420250002143, 420250002144, 420250002145,
                  420250002146, 420250002147, 420250002148, 420250002149, 420250002150, 420250002151, 420250002152, 420250002153, 420250002154,
                  420250002155, 420250002156, 420250002157, 420250002158, 420250002159, 420250002160, 420250002161, 420250002162, 420250002163,
                  420250002164, 420250002165, 420250002166, 420250002167, 420250002168, 420250002169, 420250002170, 420250002171, 420250002172,
                  420250002173, 420250002174, 420250002175, 420250002176, 420250002177, 420250002178, 420250002179, 420250002180, 420250002181,
                  420250002182, 420250002183, 420250002184, 420250002185, 420250002186, 420250002187, 420250002188, 420250002189, 420250002190,
                  420250002191, 420250002192, 420250002193, 420250002194, 420250002195, 420250002196, 420250002197, 420250002198, 420250002199,
                  420250002200, 420250002201, 420250002202, 420250002203, 420250002204, 420250002205, 420250002206, 420250002207, 420250002208,
                  420250002209, 420250002210, 420250002211, 420250002212, 420250002213, 420250002214, 420250002215, 420250002216, 420250002217,
                  420250002218, 420250002219, 420250002220, 420250002221, 420250002222, 420250002223, 420250002224, 420250002225, 420250002226,
                  420250002227, 420250002228, 420250002229, 420250002230, 420250002231, 420250002232, 420250002233, 420250002234, 420250002235,
                  420250002236, 420250002237, 420250002238, 420250002239, 420250002240, 420250002241, 420250002242, 420250002243, 420250002244,
                  420250002245, 420250002246, 420250002247, 420250002248, 420250002249, 420250002256, 420250002257, 420250002258, 420250002259,
                  420250002260, 420250002261, 420250002262, 420250002265, 420250002267, 420250002268, 420250002269, 420250002270, 420250002271,
                  420250002272, 420250002273, 420250002274]  # dtype should be int
list_of_UOM = ['10', '11', '13', '14', '15', '16', '17', '18', '19', '1A', '1B', '1C', '1D', '1E', '1F', '1G', '1H', '1I', '1J', '1K', '1L', '1M', '1X', '20',
               '21', '22', '23', '24', '25', '26', '27', '28', '29', '2A', '2B', '2C', '2I', '2J', '2K', '2L', '2M', '2N', '2P', '2Q', '2R', '2U', '2V', '2W', '2X', '2Y',
               '2Z', '30', '31', '32', '33', '34', '35', '36', '37', '38', '3B', '3C', '3E', '3G', '3H', '3I', '40', '41', '43', '44', '45', '46', '47', '48', '4A', '4B',
               '4C', '4E', '4G', '4H', '4K', '4L', '4M', '4N', '4O', '4P', '4Q', '4R', '4T', '4U', '4W', '4X', '5', '53', '54', '56', '57', '58', '59', '5A', '5B', '5C',
               '5E', '5F', '5G', '5H', '5I', '5J', '5K', '5P', '5Q', '6', '60', '61', '62', '63', '64', '66', '69', '71', '72', '73', '74', '76', '77', '78', '8', '80',
               '81', '84', '85', '87', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', 'A1', 'A10', 'A11', 'A12', 'A13', 'A14', 'A15', 'A16', 'A17', 'A18', 'A19',
               'A2', 'A20', 'A21', 'A22', 'A23', 'A24', 'A25', 'A26', 'A27', 'A28', 'A29', 'A3', 'A30', 'A31', 'A32', 'A33', 'A34', 'A35', 'A36', 'A37', 'A38', 'A39', 'A4',
               'A40', 'A41', 'A42', 'A43', 'A44', 'A45', 'A47', 'A48', 'A49', 'A5', 'A50', 'A51', 'A52', 'A53', 'A54', 'A55', 'A56', 'A57', 'A58', 'A6', 'A60', 'A61', 'A62',
               'A63', 'A64', 'A65', 'A66', 'A67', 'A68', 'A69', 'A7', 'A70', 'A71', 'A73', 'A74', 'A75', 'A76', 'A77', 'A78', 'A79', 'A8', 'A80', 'A81', 'A82', 'A83',
               'A84', 'A85', 'A86', 'A87', 'A88', 'A89', 'A9', 'A90', 'A91', 'A93', 'A94', 'A95', 'A96', 'A97', 'A98', 'AA', 'AB', 'ACR', 'AD', 'AE', 'AH', 'AI', 'AJ',
               'AK', 'AL', 'AM', 'AMH', 'AMP', 'ANN', 'AP', 'APZ', 'AQ', 'AR', 'ARE', 'AS', 'ASM', 'ASU', 'ATM', 'ATT', 'AV', 'AW', 'AY', 'AZ', 'B0', 'B1', 'B11',
               'B12', 'B13', 'B14', 'B15', 'B16', 'B18', 'B2', 'B20', 'B21', 'B22', 'B23', 'B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B3', 'B31', 'B32', 'B33', 'B34',
               'B35', 'B36', 'B37', 'B38', 'B39', 'B4', 'B40', 'B41', 'B42', 'B43', 'B44', 'B45', 'B46', 'B47', 'B48', 'B49', 'B5', 'B50', 'B51', 'B52', 'B53', 'B54',
               'B55', 'B56', 'B57', 'B58', 'B59', 'B6', 'B60', 'B61', 'B62', 'B63', 'B64', 'B65', 'B66', 'B67', 'B69', 'B7', 'B70', 'B71', 'B72', 'B73', 'B74', 'B75',
               'B76', 'B77', 'B78', 'B79', 'B8', 'B81', 'B83', 'B84', 'B85', 'B86', 'B87', 'B88', 'B89', 'B9', 'B90', 'B91', 'B92', 'B93', 'B94', 'B95', 'B96', 'B97',
               'B98', 'B99', 'BA', 'BAR', 'BB', 'BD', 'BE', 'BFT', 'BG', 'BH', 'BHP', 'BIL', 'BIM', 'BJ', 'BK', 'BL', 'BLD', 'BLL', 'BO', 'BP', 'BQL', 'BR', 'BT',
               'BTL', 'BTU', 'BUA', 'BUI', 'BW', 'BX', 'BZ', 'C0', 'C1', 'C10', 'C11', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C2', 'C20', 'C22',
               'C23', 'C24', 'C25', 'C26', 'C27', 'C28', 'C29', 'C3', 'C30', 'C31', 'C32', 'C33', 'C34', 'C35', 'C36', 'C38', 'C39', 'C4', 'C40', 'C41', 'C42',
               'C43', 'C44', 'C45', 'C46', 'C47', 'C48', 'C49', 'C5', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C59', 'C6', 'C60', 'C61',
               'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C7', 'C70', 'C71', 'C72', 'C73', 'C75', 'C76', 'C77', 'C78', 'C8', 'C80', 'C81', 'C82',
               'C83', 'C84', 'C85', 'C86', 'C87', 'C88', 'C89', 'C9', 'C90', 'C91', 'C92', 'C93', 'C94', 'C95', 'C96', 'C97', 'C98', 'C99', 'CA', 'CCT', 'CDL',
               'CEL', 'CEN', 'CG', 'CGM', 'CH', 'CJ', 'CK', 'CKG', 'CL', 'CLF', 'CLT', 'CMK', 'CMQ', 'CMT', 'CN', 'CNP', 'CNT', 'CO', 'COM', 'COU', 'CPL', 'CQ',
               'CR', 'CS', 'CT', 'CTM', 'CU', 'CUR', 'CV', 'CWA', 'CWI', 'CY', 'CZ', 'D1', 'D10', 'D12', 'D13', 'D14', 'D15', 'D16', 'D17', 'D18', 'D19',
               'D2', 'D20', 'D21', 'D22', 'D23', 'D24', 'D25', 'D26', 'D27', 'D28', 'D29', 'D30', 'D31', 'D32', 'D33', 'D34', 'D35', 'D37', 'D38', 'D39',
               'D40', 'D41', 'D42', 'D43', 'D44', 'D45', 'D46', 'D47', 'D48', 'D49', 'D5', 'D50', 'D51', 'D52', 'D53', 'D54', 'D55', 'D56', 'D57',
               'D58', 'D59', 'D6', 'D60', 'D61', 'D62', 'D63', 'D64', 'D65', 'D66', 'D67', 'D69', 'D7', 'D70', 'D71', 'D72', 'D73', 'D74', 'D75',
               'D76', 'D77', 'D79', 'D8', 'D80', 'D81', 'D82', 'D83', 'D85', 'D86', 'D87', 'D88', 'D89', 'D9', 'D90', 'D91', 'D92', 'D93', 'D94', 'D95',
               'D96', 'D97', 'D98', 'D99', 'DAA', 'DAD', 'DAY', 'DB', 'DC', 'DD', 'DE', 'DEC', 'DG', 'DI', 'DJ', 'DLT', 'DMK', 'DMQ', 'DMT', 'DN',
               'DPC', 'DPR', 'DPT', 'DQ', 'DR', 'DRA', 'DRI', 'DRL', 'DRM', 'DS', 'DT', 'DTN', 'DU', 'DWT', 'DX', 'DY', 'DZ', 'DZN', 'DZP', 'E2',
               'E3', 'E4', 'E5', 'EA', 'EB', 'EC', 'EP', 'EQ', 'EV', 'F1', 'F9', 'FAH', 'FAR', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FL',
               'FM', 'FOT', 'FP', 'FR', 'FS', 'FTK', 'FTQ', 'G2', 'G3', 'G7', 'GA', 'GB', 'GBQ', 'GC', 'GD', 'GE', 'GF', 'GFI', 'GGR', 'GH', 'GIA',
               'GII', 'GJ', 'GK', 'GL', 'GLD', 'GLI', 'GLL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GRM', 'GRN', 'GRO', 'GRT', 'GT', 'GV', 'GW', 'GWH',
               'GY', 'GZ', 'H1', 'H2', 'HA', 'HAR', 'HBA', 'HBX', 'HC', 'HD', 'HE', 'HF', 'HGM', 'HH', 'HI', 'HIU', 'HJ', 'HK', 'HL', 'HLT',
               'HM', 'HMQ', 'HMT', 'HN', 'HO', 'HP', 'HPA', 'HS', 'HT', 'HTZ', 'HUR', 'HY', 'IA', 'IC', 'IE', 'IF', 'II', 'IL', 'IM', 'INH',
               'INK', 'INQ', 'IP', 'IT', 'IU', 'IV', 'J2', 'JB', 'JE', 'JG', 'JK', 'JM', 'JO', 'JOU', 'JR', 'K1', 'K2', 'K3', 'K5', 'K6',
               'KA', 'KB', 'KBA', 'KD', 'KEL', 'KF', 'KG', 'KGM', 'KGS', 'KHZ', 'KI', 'KIT', 'KJ', 'KJO', 'KL', 'KMH', 'KMK', 'KMQ', 'KMT',
               'KNI', 'KNS', 'KNT', 'KO', 'KPA', 'KPH', 'KPO', 'KPP', 'KR', 'KS', 'KSD', 'KSH', 'KT', 'KTN', 'KUR', 'KVA', 'KVR', 'KVT',
               'KW', 'KWH', 'KWT', 'KX', 'L2', 'LA', 'LBR', 'LBT', 'LC', 'LD', 'LE', 'LEF', 'LF', 'LH', 'LI', 'LJ', 'LK', 'LM', 'LN', 'LO',
               'LP', 'LPA', 'LR', 'LS', 'LTN', 'LTR', 'LUM', 'LUX', 'LX', 'LY', 'M0', 'M1', 'M4', 'M5', 'M7', 'M9', 'MA', 'MAL', 'MAM',
                 'MAW', 'MBE', 'MBF', 'MBR', 'MC', 'MCU', 'MD', 'MF', 'MGM', 'MHZ', 'MIK', 'MIL', 'MIN', 'MIO', 'MIU', 'MK', 'MLD', 'MLT',
               'MMK', 'MMQ', 'MMT', 'MON', 'MPA', 'MQ', 'MQH', 'MQS', 'MSK', 'MT', 'MTK', 'MTQ', 'MTR', 'MTS', 'MV', 'MVA', 'MWH', 'N1',
               'N2', 'N3', 'NA', 'NAR', 'NB', 'NBB', 'NC', 'NCL', 'ND', 'NE', 'NEW', 'NF', 'NG', 'NH', 'NI', 'NIU', 'NJ', 'NL', 'NMI',
               'NMP', 'NN', 'NPL', 'NPR', 'NPT', 'NQ', 'NR', 'NRL', 'NT', 'NTT', 'NU', 'NV', 'NX', 'NY', 'OA', 'OHM', 'ON', 'ONZ', 'OP',
                'OT', 'OZ', 'OZA', 'OZI', 'P0', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'P7', 'P8', 'P9', 'PA', 'PAL', 'PB', 'PD', 'PE', 'PF',
                'PG', 'PGL', 'PHKG', 'PI', 'PK', 'PL', 'PM', 'PN', 'PO', 'PQ', 'PR', 'PS', 'PT', 'PTD', 'PTI', 'PTKG', 'PTL', 'PU',
                'PV', 'PW', 'PY', 'PZ', 'Q3', 'QA', 'QAN', 'QB', 'QD', 'QH', 'QK', 'QR', 'QT', 'QTD', 'QTI', 'QTL', 'QTR', 'R1', 'R4',
                'R9', 'RA', 'RD', 'RG', 'RH', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RPM', 'RPS', 'RS', 'RT', 'RU', 'S3', 'S4', 'S5',
                'S6', 'S7', 'S8', 'SA', 'SAN', 'SCO', 'SCR', 'SD', 'SE', 'SEC', 'SET', 'SG', 'SHT', 'SIE', 'SK', 'SL', 'SMI', 'SN', 'SO',
                 'SP', 'SQ', 'SR', 'SS', 'SST', 'ST', 'STI', 'STN', 'SV', 'SW', 'SX', 'T0', 'T1', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8',
                 'TA', 'TAH', 'TB', 'TC', 'TD', 'TE', 'TF', 'TI', 'TJ', 'TK', 'TL', 'TN', 'TNE', 'TP', 'TPR', 'TQ', 'TQD', 'TR', 'TRL',
                 'TS', 'TSD', 'TSH', 'TT', 'TU', 'TV', 'TW', 'TY', 'U1', 'U2', 'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UH', 'UM', 'VA',
                 'VI', 'VLT', 'VQ', 'VS', 'W2', 'W4', 'WA', 'WB', 'WCD', 'WE', 'WEB', 'WEE', 'WG', 'WH', 'WHR', 'WI', 'WM', 'WR', 'WSD',
                 'WTT', 'WW', 'X1', 'YDK', 'YDQ', 'YL', 'YRD', 'YT', 'Z1', 'Z2', 'Z3', 'Z4', 'Z5', 'Z6', 'Z8', 'ZP', 'ZZ']


for k in system_data_sheet.iter_rows(min_row=2):
    list_of_SIN_from_supplier_data_sheet.append(str(k[4].value))
# -----------------Making some list object for future use ----------------------------


# *********************"A.Operation" column data validation begins from here.*************************
for i in supplier_data_sheet.iter_rows(min_row=2):
    # print(i[4].value)
    Operation = str(i[0].value).strip()
    if (Operation == "delete" or Operation == "DELETE" or Operation == "Delete"):  # [0] stands for Operation
        i[0].value = "UPDATE" # capital UPDATE to identify the delete operation later
        i[0].fill = Pattern_purple
        i[9].value = "No"
        i[9].fill = Pattern_purple

    # take care of create operation
    elif (Operation == "create" or Operation == "CREATE" or Operation == "Create"):
        i[9].value = "Yes"
        i[9].fill = Pattern_purple
        i[0].fill = Pattern_purple  # just to indicate the this cell is validated

        if str(i[4].value).strip() in list_of_SIN_from_supplier_data_sheet:
            i[0].value = "update" # smallcase update to identify the operation conversion from create to update
            i[0].fill = Pattern_purple

    # take care of update operation
    elif (Operation == "update" or Operation == "UPDATE" or Operation == "Update"):
        if str(i[4].value).strip() not in list_of_SIN_from_supplier_data_sheet:
            i[0].value = "create" #smallcase create to identify the operation conversion from update to create
            i[0].fill = Pattern_purple

    elif ( Operation == "None" or len(Operation) == 4 ):  # take care of NULL operation
        i[0].value = "Not in use"
        i[0].fill = Pattern_red
    else:
        logging.warning(f"{i[4].value} , Operation column value out of scope of program")    
        i[0].fill = Pattern_warning
# ----------------------"A.Operation" column data validation Ends here.---------------------------------


# *********************"Rest all column data validation begins from here.*************************
for i in supplier_data_sheet.iter_rows(min_row=2):
    if (i[0].value == "update" or i[0].value == "UPDATE" or i[0].value == "Update"):
        # print(i[4].row)
        SIN = str(i[4].value).strip()
        # print(SIN)
        row_num_from_supplier_data_sheet = i[4].row
        # print(row_num_from_supplier_data_sheet)

        if supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=3).value is None:
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=3).value = "Material"  # Type
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=3).fill = Pattern_purple

        # Buyer Item Number, removing buyer item num
        supplier_data_sheet.cell(
            row=row_num_from_supplier_data_sheet, column=4).value = None
        supplier_data_sheet.cell(
            row=row_num_from_supplier_data_sheet, column=4).fill = Pattern_purple

        for j in system_data_sheet.iter_rows(min_row=2):

            if str(j[4].value) == SIN:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2).value = j[1].value  # Line Number*
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2).fill = Pattern_purple

                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=6).value is None):  # Short Name*
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=6).value = j[5].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=6).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=6).value = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=6).value)[0:40]  # Limiting 40 char in short name
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=6).fill = Pattern_purple

                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=7).value is None):  # Item Description*
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=7).value = j[6].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=7).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=7).value = str(supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=7).value)[0:1000]  # Limiting 1000 char in desc
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=7).fill = Pattern_purple

                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=8).value is None or
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=8).value not in list_of_UNSPSC):  # UNSPSC**
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=8).value = j[7].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=8).fill = Pattern_purple
                elif (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=8).value in list_of_UNSPSC):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=8).fill = Pattern_purple

                # Category ID** , removal of data
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=9).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=9).fill = Pattern_purple

                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=11).value is None):  # Keywords
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=11).value = j[10].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=11).fill = Pattern_purple
                else:

                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=11).value = str(
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=11).value)[0:400]  # Limiting 400 char
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=11).fill = Pattern_purple

                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=12).value is None or
                        not ((supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=12).value).isdigit())):  # Lead time
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=12).value = j[11].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=12).fill = Pattern_purple
                elif ((supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=12).value).isdigit()):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=12).fill = Pattern_purple
                else:
                    pass

                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=13).value = "USD"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=13).fill = Pattern_purple     # Currency Code* should be USD always

                # *************Price validation starts here***************
                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value is None) and (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=10).value == "No") and (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=1).value == "UPDATE"):  # price*
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14).value = j[13].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_purple
                elif type(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value) == float:
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_purple
                elif (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value is None) and (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=1).value == "Update" or supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=1).value == "update"):
                    logging.warning(
                        f"Price against SIN = {SIN} is not provided, please check {i[13]}")
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
                elif (str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=14).value).isalpha()):
                    logging.warning(
                        f"Price against SIN = {SIN} not in valid datatype, please check {i[13]}")
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14).fill = Pattern_warning
                # -------------price validation ends here---------------------


                if (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=15).value is None or
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=15).value not in list_of_UOM):  # UOM*
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=15).value = j[14].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=15).fill = Pattern_purple
                elif (supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=15).value in list_of_UOM):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=15).fill = Pattern_purple



                #********Supported UOM validation starts here*********************
                Supported_UOM = supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=18).value 
                Conversion_Factors = supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=19).value

                if Supported_UOM in list_of_UOM and (Conversion_Factors is not None or str(Conversion_Factors).isspace()):
                    if (type(Conversion_Factors) == float or type(Conversion_Factors) == int) or ( str(Conversion_Factors).isdecimal() or str(Conversion_Factors).isdigit()):
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=18).fill = Pattern_purple
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=19).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=18).value = None
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=19).value = None  
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=18).value = None
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=19).value = None  
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=18).fill = Pattern_purple
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=19).fill = Pattern_purple
                #--------Supported UOM validation ends here-----------------------


                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=20).value = None #Price per UOM
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=20).fill = Pattern_purple


                if ((supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=21).value is None) or
                    (str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=21).value)).isspace()):  # Manufacturer
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=21).value = j[20].value
                    supplier_data_sheet.cell( 
                        row=row_num_from_supplier_data_sheet, column=21).fill = Pattern_purple
                else:
                    pass


                if ((supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=22).value is None) or
                    (str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=22).value)).isspace()):  # Manufacturer Part Number
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=22).value = j[21].value
                    supplier_data_sheet.cell( 
                        row=row_num_from_supplier_data_sheet, column=22).fill = Pattern_purple
                else:
                    pass

                if ((supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=23).value is None) or
                    (str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=23).value)).isspace()):  #Manufacturer Model Number
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=23).value = j[22].value
                    supplier_data_sheet.cell( 
                        row=row_num_from_supplier_data_sheet, column=23).fill = Pattern_purple
                else:
                    pass

                #**********Minimum Order Quantity validation starts here*************         
                MinOQ = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=24).value).strip()
                """print(MinOQ, len(MinOQ) , MinOQ == "None")
                print(i[23]) -- some exercise to understand system behavoir"""

                if ( (MinOQ) == "None" or (len(MinOQ) == 0) or (MinOQ.isdigit()) ):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=24).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=24).value = None
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=24).fill = Pattern_purple
                    #logging.warning(f"MinOQ against SIN= {SIN} seems wrong datatype on cell = {i[23]}") 
                #----------Minimum Order Quantity validation end here------------------


                #**********Maximum Order Quantity validation starts here*************         
                MaxOQ = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=25).value).strip()

                if ( (MaxOQ) == "None" or (len(MaxOQ) == 0) or (MaxOQ.isdigit()) ):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=25).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=25).value = None
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=25).fill = Pattern_purple
                    #logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[24]}") 
                #----------------Maximum Order Quantity validation end here----------

                #**********Banding validation starts here*************         
                Banding = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=26).value).strip()

                if ( (Banding) == "None" or (len(Banding) == 0) or (Banding.isdigit()) ):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=26).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=26).value = None
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=26).fill = Pattern_purple
                    #logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[25]}") 
                #--------------Banding validation end here-----------------


                #*********Is Tax Exempt validation starts here************
                Is_Tax_Exempt = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=27).value).strip()   
                YesOrNo = ["Yes","YES","yes","No","NO","no"]
                if ( (Is_Tax_Exempt) == "None" or (len(Is_Tax_Exempt) == 0) or (Is_Tax_Exempt in YesOrNo) ):    
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=27).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=27).value = None    
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=27).fill = Pattern_purple
                #-------------Is Tax Exempt validation ends here----------

                #*********Contract Number validation starts here************
                Contract_Number = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=28).value).strip()
                if (  (Contract_Number == "None")  or (len(Contract_Number) <= 200)):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=28).fill = Pattern_purple
                else: # lets limit it to 200 char
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=28).value = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=28))[0:200]
                #-------------Contract Number validation ends here----------

                #*********Contract Line Number validation starts here************
                Contract_Line_Number = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=29).value).strip()
                if (  (Contract_Line_Number == "None")  or (len(Contract_Line_Number) == 0) or (Contract_Line_Number.isdigit()) ):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=29).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=29).value = None
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=29).fill = Pattern_purple    
                #-------------Contract Line Number validation ends here----------



                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=30).value = j[29].value  # Start date
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=30).fill = Pattern_purple

                #*********End date validation starts here************    
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=31).value = j[30].value  # End date
                supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=31).fill = Pattern_purple
                #-------------End date validation ends here----------

                #*********GTIN validation starts here************
                GTIN = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=32).value).strip()
                if (  (GTIN == "None")  or (len(GTIN) <= 40)):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=32).fill = Pattern_purple
                else: # lets limit it to 200 char
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=32).value = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=32))[0:40]
                #-------------GTIN validation ends here----------        


                #*********Image URL validation starts here************  
                Image_URL = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=34).value).strip()
                if ( ( Image_URL == "None" or len(Image_URL) == 0  ) and j[33].value is not None):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=34).value = j[33].value
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_purple
                elif ( ( Image_URL != "None" or len(Image_URL) != 0  )):
                    #validate the URl
                    if ( ( Image_URL[-4::] == ".jpg" ) or (Image_URL[-5::] == ".jpeg") or (Image_URL[-4::] == ".JPG") or (Image_URL[-5::] == ".JPEG")):
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_purple 
                    else:
                        logging.warning(f"URL of SIN = {SIN} at invalid , pls check {i[33]}")   
                        supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_warning  
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=34).fill = Pattern_warning

                #-------------Image URL validation ends here----------          


                #*********Green product validation starts here************
                Green_product = str(supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).value).strip()
                
                if ( ( Green_product == "None" or len(Green_product) == 0 or Green_product not in YesOrNo ) and j[34].value is not None):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).value = j[34].value
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                elif (Green_product == "Green" or Green_product == "GREEN" or Green_product == "Yes"  or Green_product == "YES"):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).value = "Yes"
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                elif (Green_product == "No"  or Green_product == "NO"):
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple
                else:
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).value = "Unknown"
                    supplier_data_sheet.cell(row=row_num_from_supplier_data_sheet, column=35).fill = Pattern_purple

                #-------------Green product validation ends here---------- 
end_time = time.time()                    
print("Execution time :",round((end_time-start_time) * 10**3,3), "ms")
supplier_data.save("save.xlsx")
