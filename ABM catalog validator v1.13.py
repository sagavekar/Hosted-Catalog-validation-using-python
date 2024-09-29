from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import datetime
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk

# from icecream import ic
from itertools import groupby
from os.path import splitext, basename


# ****************** Load excel file into variable starts here *********************************
def full_validation():
    start_time = datetime.datetime.now()
    supplier_data = load_workbook(file_path[1])  # workbookname
    # workbookname ,filepath will come from GUI
    system_data = load_workbook(file_path[0])

    system_data_sheet = system_data["Catalog Lines"]  # sheet name

    sheet_f = ""  # temp sheet name variable
    if "Catalog Lines" in supplier_data.sheetnames:
        supplier_data_sheet = supplier_data["Catalog Lines"]
    elif "catalog lines" in supplier_data.sheetnames:
        supplier_data_sheet = supplier_data["catalog lines"]

    elif len(supplier_data.sheetnames) > 1:
        for ii in supplier_data.sheetnames:
            temp_c = []
            for jj in range(1, 9):
                temp_c.append(supplier_data[ii].cell(row=1, column=jj).value)
            if (
                "Supplier Item Number*"
                or "Supplier Item Number"
                or "UNSPSC"
                or "unspsc" in temp_c
            ):
                sheet_f = ii
        supplier_data_sheet = supplier_data[sheet_f]
    else:
        supplier_data_sheet = supplier_data.active

    # ------------------- Load excel file into variable ends here --------------

    # ******************Defining colors starts here *********************************
    Pattern_purple = PatternFill(patternType="solid", fgColor="ADD8E6")
    Pattern_red = PatternFill(patternType="solid", fgColor="FFCCCB")
    Pattern_warning = PatternFill(patternType="solid", fgColor="FF0000")
    Pattern_green = PatternFill(patternType="solid", fgColor="00ff00")
    # -----------------Defining colors end here ---------------------------

    # *****************Making some list / variable object for future use*******************

    df1 = pd.read_excel(file_path[0], sheet_name="Catalog Lines")
    df1["End Date"] = pd.to_datetime(df1["End Date"])
    df1_sorted_desc = df1.sort_values("End Date", ascending=False)
    farthest_date_row = df1_sorted_desc.iloc[0]
    farthest_date = farthest_date_row["End Date"]
    farthest_date_str = farthest_date.strftime("%m/%d/%Y")
    Catalog_end_date = farthest_date_str

    """ This code is to find catalog start date -- Redundant 
    df2 = pd.read_excel(file_path[0], sheet_name="Catalog Lines")
    df2["Start Date*"] = pd.to_datetime(df2['Start Date*'])
    df2_sorted_asc = df2.sort_values('Start Date*', ascending=True)
    closest_date_row = df2_sorted_asc.iloc[0]
    closest_date = closest_date_row['Start Date*']
    closest_date_str = closest_date.strftime('%m/%d/%Y')
    Catalog_start_date = closest_date_str """

    d1 = datetime.date.today().strftime("%m/%d/%Y")  # today's date

    line_num_in_existing_catalog = []
    for l in system_data_sheet.iter_rows(min_row=2):
        line_num_in_existing_catalog.append(int(l[1].value))

    line_number_series = list(
        range(min(line_num_in_existing_catalog), max(line_num_in_existing_catalog) + 1)
    )
    difference_line_number = sorted(
        list(set(line_number_series) - set(line_num_in_existing_catalog))
    )

    # list_of_row = []  # not yet consumed

    YesOrNo = ("Yes", "YES", "yes", "No", "NO", "no")

    # Supported_image_formates = [".JPEG",".jpeg",".JPG",".jpg"]
    list_of_UNSPSC = (
        "420250001930",
        "420250001931",
        "420250001932",
        "420250001933",
        "420250001934",
        "420250001935",
        "420250002274",
        "420250001936",
        "420250001937",
        "420250001938",
        "420250001939",
        "420250001940",
        "420250001941",
        "420250001942",
        "420250001943",
        "420250001944",
        "420250001945",
        "420250001946",
        "420250001947",
        "420250001948",
        "420250001949",
        "420250001950",
        "420250001951",
        "420250001952",
        "420250001953",
        "420250001954",
        "420250001955",
        "420250001956",
        "420250001957",
        "420250001958",
        "420250001959",
        "420250001960",
        "420250001961",
        "420250001962",
        "420250001963",
        "420250001964",
        "420250001965",
        "420250001966",
        "420250001967",
        "420250001968",
        "420250001969",
        "420250001970",
        "420250001971",
        "420250001972",
        "420250001973",
        "420250001974",
        "420250001975",
        "420250001976",
        "420250001977",
        "420250001978",
        "420250001979",
        "420250001980",
        "420250001981",
        "420250001982",
        "420250001983",
        "420250001984",
        "420250001985",
        "420250001986",
        "420250001987",
        "420250001988",
        "420250001989",
        "420250001990",
        "420250001991",
        "420250001992",
        "420250001993",
        "420250001994",
        "420250001995",
        "420250001996",
        "420250001997",
        "420250001998",
        "420250001999",
        "420250002000",
        "420250002001",
        "420250002002",
        "420250002003",
        "420250002004",
        "420250002005",
        "420250002006",
        "420250002007",
        "420250002008",
        "420250002009",
        "420250002010",
        "420250002011",
        "420250002012",
        "420250002013",
        "420250002014",
        "420250002015",
        "420250002016",
        "420250002017",
        "420250002018",
        "420250002019",
        "420250002020",
        "420250002021",
        "420250002022",
        "420250002023",
        "420250002024",
        "420250002025",
        "420250002026",
        "420250002027",
        "420250002028",
        "420250002029",
        "420250002030",
        "420250002031",
        "420250002032",
        "420250002033",
        "420250002034",
        "420250002035",
        "420250002036",
        "420250002037",
        "420250002038",
        "420250002039",
        "420250002040",
        "420250002041",
        "420250002042",
        "420250002043",
        "420250002044",
        "420250002045",
        "420250002046",
        "420250002047",
        "420250002048",
        "420250002049",
        "420250002050",
        "420250002051",
        "420250002052",
        "420250002053",
        "420250002054",
        "420250002055",
        "420250002056",
        "420250002057",
        "420250002058",
        "420250002059",
        "420250002060",
        "420250002061",
        "420250002062",
        "420250002063",
        "420250002064",
        "420250002065",
        "420250002066",
        "420250002067",
        "420250002068",
        "420250002069",
        "420250002070",
        "420250002071",
        "420250002072",
        "420250002073",
        "420250002074",
        "420250002075",
        "420250002076",
        "420250002077",
        "420250002078",
        "420250002079",
        "420250002080",
        "420250002081",
        "420250002082",
        "420250002083",
        "420250002084",
        "420250002085",
        "420250002086",
        "420250002087",
        "420250002088",
        "420250002089",
        "420250002090",
        "420250002091",
        "420250002092",
        "420250002093",
        "420250002094",
        "420250002095",
        "420250002096",
        "420250002097",
        "420250002098",
        "420250002099",
        "420250002100",
        "420250002101",
        "420250002102",
        "420250002103",
        "420250002104",
        "420250002105",
        "420250002106",
        "420250002107",
        "420250002108",
        "420250002109",
        "420250002110",
        "420250002111",
        "420250002112",
        "420250002113",
        "420250002114",
        "420250002115",
        "420250002116",
        "420250002117",
        "420250002118",
        "420250002119",
        "420250002120",
        "420250002121",
        "420250002122",
        "420250002123",
        "420250002124",
        "420250002125",
        "420250002126",
        "420250002127",
        "420250002128",
        "420250002129",
        "420250002130",
        "420250002131",
        "420250002132",
        "420250002133",
        "420250002134",
        "420250002135",
        "420250002136",
        "420250002137",
        "420250002138",
        "420250002139",
        "420250002140",
        "420250002141",
        "420250002142",
        "420250002143",
        "420250002144",
        "420250002145",
        "420250002146",
        "420250002147",
        "420250002148",
        "420250002149",
        "420250002150",
        "420250002151",
        "420250002152",
        "420250002153",
        "420250002154",
        "420250002155",
        "420250002156",
        "420250002157",
        "420250002158",
        "420250002159",
        "420250002160",
        "420250002161",
        "420250002162",
        "420250002163",
        "420250002164",
        "420250002165",
        "420250002166",
        "420250002167",
        "420250002168",
        "420250002169",
        "420250002170",
        "420250002171",
        "420250002172",
        "420250002173",
        "420250002174",
        "420250002175",
        "420250002176",
        "420250002177",
        "420250002178",
        "420250002179",
        "420250002180",
        "420250002181",
        "420250002182",
        "420250002183",
        "420250002184",
        "420250002185",
        "420250002186",
        "420250002187",
        "420250002188",
        "420250002189",
        "420250002190",
        "420250002191",
        "420250002192",
        "420250002193",
        "420250002194",
        "420250002195",
        "420250002196",
        "420250002197",
        "420250002198",
        "420250002199",
        "420250002200",
        "420250002201",
        "420250002202",
        "420250002203",
        "420250002204",
        "420250002205",
        "420250002206",
        "420250002207",
        "420250002208",
        "420250002209",
        "420250002210",
        "420250002211",
        "420250002212",
        "420250002213",
        "420250002214",
        "420250002215",
        "420250002216",
        "420250002217",
        "420250002218",
        "420250002219",
        "420250002220",
        "420250002221",
        "420250002222",
        "420250002223",
        "420250002224",
        "420250002225",
        "420250002226",
        "420250002227",
        "420250002228",
        "420250002229",
        "420250002230",
        "420250002231",
        "420250002232",
        "420250002233",
        "420250002234",
        "420250002235",
        "420250002236",
        "420250002237",
        "420250002238",
        "420250002239",
        "420250002240",
        "420250002241",
        "420250002242",
        "420250002243",
        "420250002244",
        "420250002245",
        "420250002246",
        "420250002247",
        "420250002248",
        "420250002249",
        "420250002256",
        "420250002257",
        "420250002258",
        "420250002259",
        "420250002260",
        "420250002261",
        "420250002262",
        "420250002265",
        "420250002267",
        "420250002268",
        "420250002269",
        "420250002270",
        "420250002271",
        "420250002272",
        "420250002273",
    )

    list_of_UOM = (
        "RT",
        "G1",
        "2L",
        "AC",
        "AS",
        "AT",
        "HA",
        "BA",
        "BC",
        "BG",
        "BL",
        "B1",
        "BT",
        "BO",
        "BX",
        "BK",
        "BD",
        "BU",
        "CN",
        "CD",
        "CR",
        "CT",
        "CA",
        "CS",
        "CL",
        "CM",
        "CC",
        "C3",
        "FC",
        "CI",
        "IC",
        "LC",
        "M3",
        "CY",
        "CP",
        "DY",
        "D2",
        "DC",
        "DL",
        "DO",
        "DZ",
        "DR",
        "EA",
        "XL",
        "FT",
        "FF",
        "$",
        "FO",
        "GL",
        "GA",
        "GM",
        "GR",
        "H1",
        "HE",
        "HT",
        "HC",
        "HL",
        "HR",
        "HQ",
        "HH",
        "HZ",
        "CF",
        "CW",
        "IG",
        "IN",
        "JT",
        "KG",
        "K1",
        "KL",
        "KM",
        "KT",
        "K2",
        "K3",
        "LF",
        "LR",
        "LT",
        "LG",
        "TL",
        "L1",
        "LS",
        "MH",
        "MT",
        "TM",
        "MK",
        "MI",
        "MG",
        "ML",
        "MM",
        "MC",
        "MN",
        "MO",
        "OZ",
        "PK",
        "PL",
        "P1",
        "PD",
        "PR",
        "PA",
        "DW",
        "%",
        "PC",
        "PT",
        "/P",
        "PO",
        "GP",
        "LP",
        "OP",
        "Z1",
        "Z2",
        "LB",
        "L3",
        "QT",
        "QU",
        "RM",
        "RD",
        "RL",
        "SK",
        "S1",
        "SE",
        "ST",
        "S2",
        "Z",
        "SH",
        "TS",
        "SX",
        "SL",
        "SP",
        "SC",
        "SF",
        "SI",
        "SM",
        "SY",
        "SB",
        "TK",
        "MF",
        "MP",
        "MW",
        "MS",
        "TN",
        "OT",
        "TC",
        "TB",
        "T1",
        "TU",
        "TW",
        "12",
        "UN",
        "WK",
        "WU",
        "YD",
        "YR",
    )

    SIN_from_system_datasheet = []
    for k in system_data_sheet.iter_rows(min_row=2):
        SIN_from_system_datasheet.append(str(k[4].value))
    # ----------------- some list object for future use ----------------------------

    Operation_count = 0  # will display this at last
    Completed_loop = 0  # will display this tk.root
    # We will collect rows to delete, since modifying the sheet while iterating can cause issues
    rows_to_delete = []

    # Function to find consecutive row ranges
    def get_consecutive_ranges(row_numbers):
        ranges = []
        for k, g in groupby(enumerate(row_numbers), lambda ix: ix[0] - ix[1]):
            group = list(g)
            ranges.append((group[0][1], group[-1][1]))
        return ranges

    # *********************"A.Operation" column data validation begins from here.*************************
    for i in supplier_data_sheet.iter_rows(min_row=2):
        Operation = str(i[0].value).strip()
        # [0] stands for Operation
        if Operation.lower() == "delete" or Operation.lower() == "remove":
            # capital UPdate to identify the delete operation later
            i[0].value = "UPdate"
            i[0].fill = Pattern_purple
            i[9].value = "No"
            i[9].fill = Pattern_purple
            Operation_count += 1  # increare operation count by 1

        # take care of create operation
        elif Operation.lower() == "create":
            i[9].value = "Yes"
            i[9].fill = Pattern_purple
            # just to indicate the this cell is validated
            i[0].fill = Pattern_purple
            Operation_count += 1  # increare operation count by 1

            if str(i[4].value).strip() in SIN_from_system_datasheet:
                # smallcase update to identify the operation conversion from create to update
                i[0].value = "update"
                i[0].fill = Pattern_purple

        # take care of update operation
        elif Operation.lower() == "update":
            Operation_count += 1  # increare operation count by 1
            if (str(i[4].value).strip() not in SIN_from_system_datasheet) and (
                "0" + str(i[4].value).strip() not in SIN_from_system_datasheet
            ):
                # smallcase create to identify the operation conversion from update to create
                i[0].value = "create"
                i[0].fill = Pattern_purple

        # take care of NULL operation
        # Collect the row number if Operation is None or empty
        elif Operation == "None" or len(Operation) == 0:
            rows_to_delete.append(i[0].row)
            # i[0].value = "Not in use"
            # i[0].fill = Pattern_red
        else:
            logging.warning(
                f"{i[4].value} , Operation column value out of scope of program"
            )
            i[0].fill = Pattern_warning
    # ----------------------"A.Operation" column data validation Ends here.---------------------------------

    # Find consecutive ranges of rows to delete in bulk
    consecutive_row_ranges = get_consecutive_ranges(rows_to_delete)

    # Delete consecutive row ranges in bulk
    for start, end in reversed(consecutive_row_ranges):
        supplier_data_sheet.delete_rows(start, end - start + 1)

    # ********************* Rest all column data validation begins from here.exclusive for "update" opration*************************

    for i in supplier_data_sheet.iter_rows(min_row=2):
        if str(i[0].value).strip().lower() == "update":

            Completed_loop += 1
            update_progress(Completed_loop, Operation_count)

            # print(i[4].row)
            SIN = str(i[4].value).strip()
            # print(SIN)
            row_num_from_supplier_data_sheet = i[4].row
            # print(row_num_from_supplier_data_sheet) start from 2

            # ************Type validation starts here***************
            Type = (
                str(
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=3
                    ).value
                )
                .strip()
                .lower()
            )
            if Type is None or Type == "None" or len(Type) == 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).value = "Material"  # Type
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).fill = Pattern_purple
            elif (
                Type == "material"
                or Type == "fixed service"
                or Type == "variable service"
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).fill = Pattern_purple
            # -----------Type validation starts here--------------

            # ************Buyer Item Number validation starts here***************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4
            ).value = None
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4
            ).fill = Pattern_purple
            # -----------Buyer Item Number validation starts here--------------

            # *******Loop on system data starts here**********
            for j in system_data_sheet.iter_rows(min_row=2):

                if str(j[4].value).strip() == SIN:  # Match found

                    # *******Fetching line num corresponding to SIN from sytem data starts here*******
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=2
                    ).value = j[
                        1
                    ].value  # Line Number*
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=2
                    ).fill = Pattern_purple
                    # --------- Fetching line num corresponding to SIN from sytem data Ends here--------

                    # ************Short Name* validation starts here***************
                    Short_name = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6
                        ).value
                    ).strip()
                    if (
                        Short_name is None
                        or Short_name == "None"
                        or len(Short_name) == 0
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6
                        ).value = j[5].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6
                        ).fill = Pattern_purple
                    else:
                        # Limiting 40 char in short name
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6
                        ).value = Short_name[0:40]
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=6
                        ).fill = Pattern_purple
                    # -------Short Name* validation ends here------------

                    # ************Item Description* validation starts here***************
                    Item_Description = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7
                        ).value
                    ).strip()
                    if (
                        Item_Description is None
                        or Item_Description == "None"
                        or len(Item_Description) == 0
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7
                        ).value = j[6].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7
                        ).fill = Pattern_purple
                    else:
                        # Limiting 1000 char in desc
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7
                        ).value = Item_Description[0:1000]
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=7
                        ).fill = Pattern_purple
                    # -------------Item Description* validation ends here------------

                    # ************ UNSPSC* validation starts here***************
                    UNSPSC = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8
                        ).value
                    ).strip()
                    if (
                        UNSPSC is None
                        or UNSPSC == "None"
                        or len(UNSPSC) == 0
                        or UNSPSC not in list_of_UNSPSC
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8
                        ).value = j[7].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8
                        ).fill = Pattern_purple
                    elif UNSPSC in list_of_UNSPSC:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=8
                        ).fill = Pattern_purple
                    # ------------- UNSPSC* validation ends here------------

                    # ************ Category ID**  validation starts here **************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=9
                    ).value = None
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=9
                    ).fill = Pattern_purple
                    # ------------- Category ID**  validation starts here -------------

                    # ************ Keywords  validation starts here **************
                    Keywords = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11
                        ).value
                    ).strip()
                    if Keywords is None or Keywords == "None" or len(Keywords) == 0:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11
                        ).value = j[10].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11
                        ).value = Keywords[
                            0:400
                        ]  # Limiting 400 char
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=11
                        ).fill = Pattern_purple
                    # ---------------- Keywords  validation starts here ---------------

                    # ************  Lead time  validation starts here **************
                    Lead_time = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12
                        ).value
                    ).strip()
                    if (
                        Lead_time is None
                        or Lead_time == "None"
                        or len(Lead_time) == 0
                        or not (Lead_time.isdigit())
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12
                        ).value = j[11].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12
                        ).fill = Pattern_purple
                    elif Lead_time.isdigit():
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=12
                        ).fill = Pattern_purple
                    else:
                        pass
                    # ---------------- Lead time validation starts here ---------------

                    # *************Currency Code* should be USD always*************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=13
                    ).value = "USD"
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=13
                    ).fill = Pattern_purple
                    # ---------------Currency Code* should be USD always-------------

                    # *************Price validation starts here***************
                    Price = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14
                        ).value
                    ).strip()
                    if (
                        (Price == "None" or Price is None or len(Price) == 0)
                        and (
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=10
                            ).value
                            == "No"
                        )
                        and (
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=1
                            ).value
                            == "UPDATE"
                        )
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14
                        ).value = j[13].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14
                        ).fill = Pattern_purple
                    elif (
                        type(
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=14
                            ).value
                        )
                        == float
                    ) or (
                        type(
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=14
                            ).value
                        )
                        == int
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14
                        ).fill = Pattern_purple
                    elif (Price == "None" or Price is None or len(Price) == 0) and (
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=1
                        ).value
                        == "Update"
                        or supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=1
                        ).value
                        == "update"
                    ):
                        logging.warning(
                            f"Price against SIN = {SIN} is not provided, please check {i[13]}"
                        )
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14
                        ).fill = Pattern_warning
                    elif Price.isalpha():
                        logging.warning(
                            f"Price against SIN = {SIN} not in valid datatype, please check {i[13]}"
                        )
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=14
                        ).fill = Pattern_warning
                    # -------------price validation ends here---------------------

                    # ************* UOM* validation starts here***************
                    UOM = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15
                        ).value
                    ).strip()
                    if UOM == "None" or len(UOM) == 0 or UOM not in list_of_UOM:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15
                        ).value = j[14].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15
                        ).fill = Pattern_purple
                    elif UOM in list_of_UOM:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=15
                        ).fill = Pattern_purple
                    # ------------- UOM* validation ends here---------------------

                    # ********Supported UOM validation starts here*********************
                    Supported_UOM = supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=18
                    ).value
                    Conversion_Factors = supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=19
                    ).value

                    if (
                        (Supported_UOM in list_of_UOM)
                        and (
                            Conversion_Factors is not None
                            or str(Conversion_Factors).isspace()
                        )
                        and (Supported_UOM != UOM)
                    ):
                        if (
                            type(Conversion_Factors) == float
                            or type(Conversion_Factors) == int
                        ) or (str(Conversion_Factors).isdigit()):
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=18
                            ).fill = Pattern_purple
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=19
                            ).fill = Pattern_purple
                            try:
                                supplier_data_sheet.cell(
                                    row=row_num_from_supplier_data_sheet, column=20
                                ).value = float(Conversion_Factors) * float(Price)
                                supplier_data_sheet.cell(
                                    row=row_num_from_supplier_data_sheet, column=20
                                ).fill = Pattern_purple
                            except:
                                supplier_data_sheet.cell(
                                    row=row_num_from_supplier_data_sheet, column=20
                                ).fill = Pattern_purple
                        else:
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=18
                            ).value = None
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=19
                            ).value = None
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=18
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=18
                        ).fill = Pattern_purple
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=19
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=19
                        ).fill = Pattern_purple
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20
                        ).fill = Pattern_purple
                    # Supported UOM validation ends here-----------------------

                    # ********Price per UOM validation starts here********
                    # this has been done in above code already
                    # --------Price per UOM validation starts here--------

                    # ********Manufacturer validation starts here********
                    Manufacturer = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=21
                        ).value
                    )

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=21
                    ).value = (Manufacturer[0:50] if Manufacturer != "None" else "")

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=21
                    ).fill = Pattern_purple

                    # --------Manufacturer validation starts here--------

                    # ********Manufacturer Part Number validation starts here********
                    Manufacturer_Part_Number = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=22
                        ).value
                    )

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=22
                    ).value = (
                        Manufacturer_Part_Number[0:256]
                        if Manufacturer_Part_Number != "None"
                        else ""
                    )

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=22
                    ).fill = Pattern_purple
                    # --------Manufacturer Part Number validation starts here--------

                    # ********Manufacturer Model Number validation starts here********
                    Manufacturer_Model_Number = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=23
                        ).value
                    )

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=23
                    ).value = (
                        Manufacturer_Model_Number[0:500]
                        if Manufacturer_Model_Number != "None"
                        else ""
                    )

                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=23
                    ).fill = Pattern_purple
                    # --------Manufacturer Model Number validation starts here--------

                    # **********Minimum Order Quantity validation starts here*************
                    MinOQ = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24
                        ).value
                    ).strip()
                    """print(MinOQ, len(MinOQ) , MinOQ == "None")
                    print(i[23]) -- some exercise to understand system behavoir"""

                    if (MinOQ) == "None" or (len(MinOQ) == 0) or (MinOQ.isdigit()):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=24
                        ).fill = Pattern_purple
                        # logging.warning(f"MinOQ against SIN= {SIN} seems wrong datatype on cell = {i[23]}")
                    # ----------Minimum Order Quantity validation end here------------------

                    # **********Maximum Order Quantity validation starts here*************
                    MaxOQ = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25
                        ).value
                    ).strip()

                    if (MaxOQ) == "None" or (len(MaxOQ) == 0) or (MaxOQ.isdigit()):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=25
                        ).fill = Pattern_purple
                        # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[24]}")
                    # ----------------Maximum Order Quantity validation end here----------

                    # **********Banding validation starts here*************
                    Banding = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26
                        ).value
                    ).strip()

                    if (
                        (Banding) == "None"
                        or (len(Banding) == 0)
                        or (Banding.isdigit())
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=26
                        ).fill = Pattern_purple
                        # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[25]}")
                    # --------------Banding validation end here-----------------

                    # *********Is Tax Exempt validation starts here************
                    Is_Tax_Exempt = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27
                        ).value
                    ).strip()

                    if (
                        (Is_Tax_Exempt) == "None"
                        or (len(Is_Tax_Exempt) == 0)
                        or (Is_Tax_Exempt in YesOrNo)
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=27
                        ).fill = Pattern_purple
                    # -------------Is Tax Exempt validation ends here----------

                    # *********Contract Number validation starts here************
                    Contract_Number = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=28
                        ).value
                    ).strip()
                    if (Contract_Number == "None") or (len(Contract_Number) <= 200):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=28
                        ).fill = Pattern_purple
                    else:  # lets limit it to 200 char
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=28
                        ).value = Contract_Number[0:200]
                        logging.warning(
                            f"{SIN}, Contract number exceding char limit, pls check {i[27]}"
                        )
                    # -------------Contract Number validation ends here----------

                    # *********Contract Line Number validation starts here************
                    Contract_Line_Number = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29
                        ).value
                    ).strip()
                    if (
                        (Contract_Line_Number == "None")
                        or (len(Contract_Line_Number) == 0)
                        or (Contract_Line_Number.isdigit())
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=29
                        ).fill = Pattern_purple
                    # -------------Contract Line Number validation ends here----------

                    # *********Start date validation starts here************
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=30
                    ).value = j[29].value
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=30
                    ).fill = Pattern_purple
                    # -------------Start date validation ends here----------

                    # *********End date validation starts here************
                    if str(i[0].value).strip() == "UPdate":
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=31
                        ).value = datetime.datetime.today().strftime("%m/%d/%Y")

                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=31
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=31
                        ).value = str(Catalog_end_date)
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=31
                        ).fill = Pattern_purple
                    # -------------End date validation ends here----------

                    # *********GTIN validation starts here************
                    GTIN = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32
                        ).value
                    ).strip()
                    if (GTIN == "None") or (len(GTIN) <= 40):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32
                        ).fill = Pattern_purple
                    else:  # lets limit it to 200 char
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32
                        ).value = GTIN[0:40]
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=32
                        ).fill = Pattern_purple
                    # -------------GTIN validation ends here----------

                    # *********Image URL validation starts here************
                    Image_URL = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34
                        ).value
                    ).strip()

                    if (Image_URL == "None" or len(Image_URL) == 0) and str(
                        j[33].value
                    ).strip() != "None":
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34
                        ).value = j[33].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34
                        ).fill = Pattern_purple
                    elif Image_URL != "None" or len(Image_URL) != 0:
                        # validate the URL formates here
                        if (
                            (Image_URL[-4::] == ".jpg")
                            or (Image_URL[-5::] == ".jpeg")
                            or (Image_URL[-4::] == ".JPG")
                            or (Image_URL[-5::] == ".JPEG")
                        ):
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=34
                            ).fill = Pattern_green
                        else:
                            logging.warning(
                                f"URL of SIN = {SIN} at invalid , pls check {i[33]}"
                            )
                            supplier_data_sheet.cell(
                                row=row_num_from_supplier_data_sheet, column=34
                            ).fill = Pattern_warning
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=34
                        ).fill = Pattern_warning

                    # -------------Image URL validation ends here----------

                    # -------------Image Name validation Start here----------

                    if Image_URL != "None":
                        # print("Image_URL is None")
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=33
                        ).value = None
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=33
                        ).fill = Pattern_purple

                    # *********Image Name validation Ends here************

                    # *********Green product validation starts here************
                    Green_product = str(
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).value
                    ).strip()

                    if (
                        Green_product == "None"
                        or len(Green_product) == 0
                        or Green_product not in YesOrNo
                    ) and j[34].value is not None:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).value = j[34].value
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).fill = Pattern_purple
                    elif (
                        Green_product == "Green"
                        or Green_product == "GREEN"
                        or Green_product == "Yes"
                        or Green_product == "YES"
                    ):
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).value = "Yes"
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).fill = Pattern_purple
                    elif Green_product == "No" or Green_product == "NO":
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).value = "No"
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).fill = Pattern_purple
                    else:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).value = "Unknown"
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=35
                        ).fill = Pattern_purple

                    # -------------Green product validation ends here----------
            # *******Loop on system data Ends here**********
    # ---------------"Rest all column data validation ends here.exclusive for "update" opration------------------------

    # ********************* Rest all column data validation begins from here.exclusive for "Create" opration*******************

    for i in supplier_data_sheet.iter_rows(min_row=2):
        if str(i[0].value).lower() == "create":

            Completed_loop += 1
            update_progress(Completed_loop, Operation_count)

            SIN = str(i[4].value).strip()
            row_num_from_supplier_data_sheet = i[4].row

            # ************Type validation starts here***************
            Type = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).value
            ).strip()
            if Type is None or Type == "None" or len(Type) == 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).value = "Material"  # Default Type
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).fill = Pattern_purple
            elif (
                Type == "material"
                or Type == "fixed service"
                or Type == "variable service"
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=3
                ).fill = Pattern_purple
            else:
                pass
            # -----------Type validation starts here--------------

            # ************Buyer Item Number validation starts here***************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4
            ).value = None
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=4
            ).fill = Pattern_purple
            # -----------Buyer Item Number validation starts here--------------

            # *******Line number validation starts here *******
            if len(difference_line_number) != 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2
                ).value = difference_line_number[0]
                # assing the value and the  pop the value
                difference_line_number.pop(0)
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2
                ).fill = Pattern_purple
            elif len(difference_line_number) == 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2
                ).value = (max(line_num_in_existing_catalog) + 1)
                line_num_in_existing_catalog.append(
                    max(line_num_in_existing_catalog) + 1
                )
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=2
                ).fill = Pattern_warning
            # --------Line number validation ends here ---------

            # ************Short Name* validation starts here***************
            Short_name = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6
                ).value
            ).strip()
            if Short_name is None or Short_name == "None" or len(Short_name) == 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6
                ).value = SIN
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6
                ).fill = Pattern_purple
            else:
                # Limiting 40 char in short name
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6
                ).value = Short_name[0:40]
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6
                ).fill = Pattern_purple
            # -------Short Name* validation ends here------------

            # ************Item Description* validation starts here***************
            Item_Description = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7
                ).value
            ).strip()
            if (
                Item_Description is None
                or Item_Description == "None"
                or len(Item_Description) == 0
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7
                ).value = supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=6
                ).value
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7
                ).fill = Pattern_purple
            else:
                # Limiting 1000 char in desc
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7
                ).value = Item_Description[0:1000]
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=7
                ).fill = Pattern_purple
            # -------------Item Description* validation ends here------------

            # ************ UNSPSC* validation starts here***************
            UNSPSC = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8
                ).value
            ).strip()
            if (
                UNSPSC is None
                or UNSPSC == "None"
                or len(UNSPSC) == 0
                or UNSPSC not in list_of_UNSPSC
            ):
                # defaulting UNSPSC code
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8
                ).value = "420250002033"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8
                ).fill = Pattern_purple
                logging.info("UNSPSC against SIN = {SIN}, defaulted to 420250002033 ")
            elif UNSPSC in list_of_UNSPSC:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=8
                ).fill = Pattern_purple
            # ------------- UNSPSC* validation ends here------------

            # ************ Category ID**  validation starts here **************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=9
            ).value = None
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=9
            ).fill = Pattern_purple
            # ------------- Category ID**  validation starts here -------------

            # ************ Keywords  validation starts here **************
            Keywords = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11
                ).value
            ).strip()
            if Keywords is None or Keywords == "None" or len(Keywords) == 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11
                ).value = Keywords[
                    0:400
                ]  # Limiting 400 char
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=11
                ).fill = Pattern_purple
            # ---------------- Keywords  validation starts here ---------------

            # ************  Lead time  validation starts here **************
            Lead_time = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12
                ).value
            ).strip()
            if (
                Lead_time is None
                or Lead_time == "None"
                or len(Keywords) == 0
                or not (Lead_time.isdigit())
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12
                ).value = 10  # defaulting lead time to 10
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12
                ).fill = Pattern_purple
                logging.info(f"LT of SIN = {SIN} defaulted to 10 at {i[11]}")
            elif Lead_time.isdigit():
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=12
                ).fill = Pattern_purple
            else:
                pass
            # ---------------- Lead time validation starts here ---------------

            # *************Currency Code* should be USD always*************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=13
            ).value = "USD"
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=13
            ).fill = Pattern_purple
            # ---------------Currency Code* should be USD always-------------

            # *************Price validation starts here***************
            Price = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14
                ).value
            ).strip()
            if Price == "None" or Price is None or len(Price) == 0:
                logging.warning(
                    f"New SIN = {SIN}, price is not provided, pls check with supplier"
                )
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14
                ).fill = Pattern_warning
            elif (
                type(
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14
                    ).value
                )
                == float
            ) or (
                type(
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=14
                    ).value
                )
                == int
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14
                ).fill = Pattern_purple
            elif Price is None or Price == "None" or len(Price) == 0:
                logging.warning(
                    f"New SIN = {SIN}, price is not provided, pls check with supplier"
                )
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14
                ).fill = Pattern_warning
            elif Price.isalpha():
                logging.warning(
                    f"Price against SIN = {SIN} have invalid datatype, please check {i[13]}"
                )
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=14
                ).fill = Pattern_warning
            # -------------price validation ends here---------------------

            # ************* UOM* validation starts here***************
            UOM = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15
                ).value
            ).strip()
            if UOM == "None" or len(UOM) == 0 or UOM not in list_of_UOM:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15
                ).value = "Invalid_UOM"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15
                ).fill = Pattern_red
                logging.info(f"UOM of SIN = {SIN} defaulted to EA")
            elif UOM in list_of_UOM:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=15
                ).fill = Pattern_purple
            # ------------- UOM* validation ends here---------------------

            # ********Supported UOM validation starts here*********************
            Supported_UOM = supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=18
            ).value
            Conversion_Factors = supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=19
            ).value

            if (
                (Supported_UOM in list_of_UOM)
                and (
                    Conversion_Factors is not None or str(Conversion_Factors).isspace()
                )
                and (Supported_UOM != UOM)
            ):
                if (
                    type(Conversion_Factors) == float or type(Conversion_Factors) == int
                ) or (
                    str(Conversion_Factors).isdecimal()
                    or str(Conversion_Factors).isdigit()
                ):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=18
                    ).fill = Pattern_purple
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=19
                    ).fill = Pattern_purple
                    try:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20
                        ).value = float(Conversion_Factors) * float(Price)
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20
                        ).fill = Pattern_purple
                    except:
                        supplier_data_sheet.cell(
                            row=row_num_from_supplier_data_sheet, column=20
                        ).fill = Pattern_purple

                else:
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=18
                    ).value = None
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=19
                    ).value = None

            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=18
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=18
                ).fill = Pattern_purple
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=19
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=19
                ).fill = Pattern_purple
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=20
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=20
                ).fill = Pattern_purple
            # Supported UOM validation ends here-----------------------

            # ********Price per UOM validation starts here********
            # Not needed since above code takes of this part
            # --------Price per UOM validation starts here--------

            # ********Manufacturer validation starts here********
            # Validation not actually applied
            Manufacturer = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=21
                ).value
            ).strip()

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=21
            ).value = Manufacturer[0:50]

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=21
            ).fill = Pattern_purple
            # --------Manufacturer validation starts here--------

            # ********Manufacturer Part Number validation starts here********
            # Validation not actually applied
            Manufacturer_Part_Number = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=22
                ).value
            ).strip()

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=22
            ).value = Manufacturer_Part_Number[0:256]

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=22
            ).fill = Pattern_purple
            # --------Manufacturer Part Number validation starts here--------

            # ********Manufacturer Model Number validation starts here********

            Manufacturer_Model_Number = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=23
                ).value
            ).strip()

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=23
            ).value = Manufacturer_Model_Number[0:500]

            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=23
            ).fill = Pattern_purple
            # --------Manufacturer Model Number validation starts here--------

            # **********Minimum Order Quantity validation starts here*************
            MinOQ = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24
                ).value
            ).strip()
            """print(MinOQ, len(MinOQ) , MinOQ == "None")
            print(i[23]) -- some exercise to understand system behavoir"""

            if (MinOQ) == "None" or (len(MinOQ) == 0) or (MinOQ.isdigit()):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=24
                ).fill = Pattern_purple
                # logging.warning(f"MinOQ against SIN= {SIN} seems wrong datatype on cell = {i[23]}")
            # ----------Minimum Order Quantity validation end here------------------

            # **********Maximum Order Quantity validation starts here*************
            MaxOQ = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25
                ).value
            ).strip()

            if (MaxOQ) == "None" or (len(MaxOQ) == 0) or (MaxOQ.isdigit()):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=25
                ).fill = Pattern_purple
                # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[24]}")
            # ----------------Maximum Order Quantity validation end here----------

            # **********Banding validation starts here*************
            Banding = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26
                ).value
            ).strip()

            if (Banding) == "None" or (len(Banding) == 0) or (Banding.isdigit()):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=26
                ).fill = Pattern_purple
                # logging.warning(f"MaxOQ against SIN= {SIN} seems wrong datatype on cell = {i[25]}")
            # --------------Banding validation end here-----------------

            # *********Is Tax Exempt validation starts here************
            Is_Tax_Exempt = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27
                ).value
            ).strip()

            if (
                (Is_Tax_Exempt) == "None"
                or (len(Is_Tax_Exempt) == 0)
                or (Is_Tax_Exempt in YesOrNo)
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=27
                ).fill = Pattern_purple
            # -------------Is Tax Exempt validation ends here----------

            # *********Contract Number validation starts here************
            Contract_Number = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=28
                ).value
            ).strip()
            if (Contract_Number == "None") or (len(Contract_Number) <= 200):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=28
                ).fill = Pattern_purple
            else:  # lets limit it to 200 char
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=28
                ).value = Contract_Number[0:200]
                logging.warning(
                    f"{SIN}, Contract number exceding char limit, pls check {i[27]}"
                )
            # -------------Contract Number validation ends here----------

            # *********Contract Line Number validation starts here************
            Contract_Line_Number = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29
                ).value
            ).strip()
            if (
                (Contract_Line_Number == "None")
                or (len(Contract_Line_Number) == 0)
                or (Contract_Line_Number.isdigit())
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=29
                ).fill = Pattern_purple
            # -------------Contract Line Number validation ends here----------

            # *********Start date validation starts here************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=30
            ).value = str(
                d1
            )  # add today's date
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=30
            ).fill = Pattern_purple
            # -------------Start date validation ends here----------

            # *********End date validation starts here************
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=31
            ).value = str(Catalog_end_date)
            supplier_data_sheet.cell(
                row=row_num_from_supplier_data_sheet, column=31
            ).fill = Pattern_purple
            # -------------End date validation ends here----------

            # *********GTIN validation starts here************
            GTIN = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32
                ).value
            ).strip()
            if (GTIN == "None") or (len(GTIN) <= 40):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32
                ).fill = Pattern_purple
            else:  # lets limit it to 200 char
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32
                ).value = GTIN[0:40]
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=32
                ).fill = Pattern_purple
            # -------------GTIN validation ends here----------

            # *********Image URL validation starts here************
            Image_URL = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=34
                ).value
            ).strip()
            if Image_URL == "None" or len(Image_URL) == 0:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=34
                ).fill = Pattern_purple
            elif Image_URL != "None" or len(Image_URL) != 0:
                # validate the URL formates here
                if (
                    (Image_URL[-4::] == ".jpg")
                    or (Image_URL[-5::] == ".jpeg")
                    or (Image_URL[-4::] == ".JPG")
                    or (Image_URL[-5::] == ".JPEG")
                ):
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=34
                    ).fill = Pattern_green
                else:
                    logging.warning(
                        f"URL of SIN = {SIN} at invalid , pls check {i[33]}"
                    )
                    supplier_data_sheet.cell(
                        row=row_num_from_supplier_data_sheet, column=34
                    ).fill = Pattern_warning
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=34
                ).fill = Pattern_warning

            # -------------Image URL validation ends here----------

            # *********Image Name validation starts here***********

            if Image_URL != "None":
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=33
                ).value = None
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=33
                ).fill = Pattern_purple
            # -------------Image Name validation ends here----------

            # *********Green product validation starts here************
            Green_product = str(
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).value
            ).strip()

            if (
                Green_product == "None"
                or len(Green_product) == 0
                or Green_product not in YesOrNo
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).value = "Unknown"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).fill = Pattern_purple
            elif (
                Green_product == "Green"
                or Green_product == "GREEN"
                or Green_product == "Yes"
                or Green_product == "YES"
                or Green_product == "Yes-Certified"
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).value = "Yes"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).fill = Pattern_purple
            elif (
                Green_product == "No"
                or Green_product == "NO"
                or Green_product == "Yes-Not Certified"
            ):
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).value = "No"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).fill = Pattern_purple
            else:
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).value = "Unknown"
                supplier_data_sheet.cell(
                    row=row_num_from_supplier_data_sheet, column=35
                ).fill = Pattern_purple

            # -------------Green product validation ends here----------
    end_time = datetime.datetime.now()
    ex_time = (
        f"Total operations = {Operation_count} & Execution time = {end_time-start_time}"
    )
    tk.messagebox.showinfo("Task completed", ex_time)

    stats = tk.Text(root, font="Tahoma 11")
    stats.insert(tk.INSERT, ex_time)
    stats.pack()

    # ---------------- Rest all column data validation ends  here.exclusive for "Create" opration-----------------

    supplier_data.save(
        f"VALIDATED_{datetime.datetime.now().strftime('%m.%d.%Y %H.%M.%S')}__{file_name2}.xlsx"
    )  # final save the file

    # *******Opening of newly created excel*******
    "No code"
    # --------Opening of newly created excel-------


# ***** GUI buidling starts here********
file_path = []


def clear_path():
    file_path.clear()
    file_label1.config(text="")  # destroy filelable1
    file_label2.config(text="")  # destroy filelable2


def browse_file1():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", ("*.xlsx", "*.xls"))]
    )
    file_label1.config(text=file_path)


def browse_file2():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", ("*.xlsx", "*.xls"))]
    )
    global file_name2  # this is for final save filename - do not delete
    file_name2 = splitext(basename(file_path))[0]
    file_label2.config(text=file_path)


def load_files():
    file_path1 = file_label1.cget("text")
    file_path2 = file_label2.cget("text")
    file_path.append(file_path1)
    file_path.append(file_path2)
    if file_path1 and file_path2:
        result = True
    else:
        tk.messagebox.showerror(
            title="Missing files", message="Please select both files first"
        )
        result = False
    return result

def update_progress(Completed_loop, Operation_count):
    progress_percentage = (Completed_loop / Operation_count) * 100
    progress["value"] = progress_percentage
    root.update_idletasks()


root = tk.Tk()
root.geometry("500x380")
root.configure(background="#F0F0F0")
root.title(
    "ABM Catalog Validator V1.13 | sagavekar.om@gmail.com",
)
root.minsize(500, 390)

label1 = tk.Label(root, text="Select System Extract", font="Tahoma 13")
label1.pack()

button1 = tk.Button(
    root,
    text="Browse",
    command=browse_file1,
    font="Tahoma 13",
    width=16,
    activebackground="blue",
    relief="groove",
    bg="#6EC5EC",
    fg="black",
)
button1.pack()

file_label1 = tk.Label(root, text="")
file_label1.pack()

label2 = tk.Label(root, text="Select Supplier Template", font="Tahoma 13")
label2.pack()

button2 = tk.Button(
    root,
    text="Browse",
    command=browse_file2,
    font="Tahoma 13",
    width=16,
    activebackground="blue",
    relief="groove",
    bg="#6EC5EC",
    fg="black",
)
button2.pack()

file_label2 = tk.Label(root, text="")
file_label2.pack()

run_button = tk.Button(
    root,
    text="Validate",
    command=lambda: (load_files(), full_validation()),
    font="Tahoma 13",
    width=16,
    activebackground="blue",
    relief="groove",
    bg="#6EC5EC",
    fg="black",
)
run_button.pack(pady=5)

reset_button = tk.Button(
    root,
    text="Reset",
    command=clear_path,
    font="Tahoma 13",
    width=16,
    activebackground="blue",
    relief="groove",
    bg="#6EC5EC",
    fg="black",
)
reset_button.pack(pady=5)

progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress.pack()


"""label3 = tk.Label(root, text="Designed and developed by Omkar Sagavekar",
                  font="Tahoma 10", bg="black", fg="white")
label3.pack(anchor="s", fill="x", side="bottom")
"""

root.mainloop()  # this for GUI loop

# ***** GUI buidling ends here********
